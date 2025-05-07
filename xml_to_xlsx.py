import openpyxl
from xml.etree import ElementTree

# Caminho do arquivo XML
xml_path = 'C:\\python\\ajuda.xml'

try:
    # Carregar e analisar o arquivo XML
    tree = ElementTree.parse(xml_path)
    root = tree.getroot()
    
    # Criar um novo arquivo Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Verificar se o XML contém elementos
    if len(root) == 0:
        raise ValueError("O arquivo XML está vazio ou mal formatado.")

    # Escrever a linha de cabeçalho
    header_row = root[0]
    for i, elem in enumerate(header_row):
        sheet.cell(row=1, column=i+1, value=elem.tag)  # Usa o nome do elemento como cabeçalho

    # Escrever as linhas de dados
    for row_idx, row in enumerate(root[1:], start=2):  # A partir da segunda linha
        for col_idx, cell in enumerate(row):
            sheet.cell(row=row_idx, column=col_idx+1, value=cell.text)

    # Salvar o arquivo Excel
    workbook.save('output.xlsx')
    print("Arquivo Excel 'output.xlsx' criado com sucesso!")

except FileNotFoundError:
    print(f"Erro: O arquivo XML '{xml_path}' não foi encontrado.")
except ElementTree.ParseError as e:
    print(f"Erro ao analisar o XML: {e}")
except Exception as e:
    print(f"Ocorreu um erro: {e}")
